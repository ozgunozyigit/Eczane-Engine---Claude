import re
import math
import os
import calendar
from io import BytesIO
from datetime import datetime, date, timedelta, timezone
from typing import Optional

import pandas as pd
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from rapidfuzz import fuzz

app = FastAPI(title="Eczane Sipariş ENGINe API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================================================
# SABİT LİSTE DIŞI BAŞLANGIÇLAR
# =========================================================

LISTE_DISI_BASLANGICLAR_RAW = [
    "BIODERMA", "CUBITAN", "DIASIP", "ENSURE", "FORTIMEL", "FORTINI",
    "GLUCERNA", "IMPACT", "LAROCHE", "NUTRISON", "NUTRIVIGOR", "PEDIASURE",
    "PEPTAMEN", "PHYSIONEAL", "RESOURCE", "ENJEKTOR", "MAJISTRAL", "SENTE",
    "SARIINTRAKET", "INFATRINI", "VITAL", "SANITAYARA", "OCEAN", "NUXE", "NBTLIFE", "NUTRAXIN",
    "SOLGAR", "VELAVIT", "BIOXCIN", "CERAVE", "DERMOSKIN", "DINAMIS", "DYNAVIT", "EASYFISH", "MEDICAGO",
    "MOLLERS", "MUSTELA", "NATURESBOUNTY", "REDEXON", "ROLL", "SENSODYNE", "SIHHAT", "TOPICREM", "TTO", "VICHY",
    "ZADEVITAL", "WELLCARE", "AYSET",
    "NEWLIFE", "STREPSILS", "CURAPROX", "INTRAKET", "ACTIVE PLUS", "ARGIVIT", "BEPANTHOL",
    "CAN PED", "CAPICADE", "CAUDALIE", "DAY 2", "DUCRAY", "ENTEROGERMINA", "GOODDAY",
    "GRIPIN", "IMUNEKS", "IMUNOL", "NBL", "NIVEA", "NTB", "NOW", "NYXON", "OKEY",
    "PARADONTAX", "REDOXON", "SKINCEUT", "SMARTUP", "SORVAGEN", "STERIMAR", "SUPRADYN",
    "PHARMATON", "TABIA", "TABVITAMIN", "TURKFLEKS", "UMCA", "VEDEXA", "VENATURA",
    "VERISCA", "WEE",
]

TURKCE_AYLAR = {
    1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan",
    5: "Mayıs", 6: "Haziran", 7: "Temmuz", 8: "Ağustos",
    9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık",
}


def bugun_turkiye():
    return (datetime.now(timezone.utc) + timedelta(hours=3)).date()


def ay_ekle(tarih, ay_sayisi):
    yeni_ay_index = tarih.month - 1 + ay_sayisi
    yeni_yil = tarih.year + yeni_ay_index // 12
    yeni_ay = yeni_ay_index % 12 + 1
    yeni_gun = min(tarih.day, calendar.monthrange(yeni_yil, yeni_ay)[1])
    return date(yeni_yil, yeni_ay, yeni_gun)


def turkce_tarih_yaz(tarih):
    return f"{tarih.day:02d} {TURKCE_AYLAR[tarih.month]} {tarih.year}"


def onerilen_rapor_araligi():
    bugun = bugun_turkiye()
    bu_ayin_ilk_gunu = date(bugun.year, bugun.month, 1)
    baslangic = ay_ekle(bu_ayin_ilk_gunu, -3)
    bitis = bu_ayin_ilk_gunu - timedelta(days=1)
    return baslangic, bitis


def turkiye_resmi_tatilleri(yil):
    tatiller = set()
    tatiller.add(date(yil, 1, 1))
    tatiller.add(date(yil, 4, 23))
    tatiller.add(date(yil, 5, 1))
    tatiller.add(date(yil, 5, 19))
    tatiller.add(date(yil, 7, 15))
    tatiller.add(date(yil, 8, 30))
    tatiller.add(date(yil, 10, 29))

    dini_tatiller = {
        2025: [date(2025,3,30),date(2025,3,31),date(2025,4,1),date(2025,6,6),date(2025,6,7),date(2025,6,8),date(2025,6,9)],
        2026: [date(2026,3,20),date(2026,3,21),date(2026,3,22),date(2026,5,27),date(2026,5,28),date(2026,5,29),date(2026,5,30)],
        2027: [date(2027,3,9),date(2027,3,10),date(2027,3,11),date(2027,5,16),date(2027,5,17),date(2027,5,18),date(2027,5,19)],
        2028: [date(2028,2,26),date(2028,2,27),date(2028,2,28),date(2028,5,4),date(2028,5,5),date(2028,5,6),date(2028,5,7)],
        2029: [date(2029,2,14),date(2029,2,15),date(2029,2,16),date(2029,4,23),date(2029,4,24),date(2029,4,25),date(2029,4,26)],
        2030: [date(2030,2,4),date(2030,2,5),date(2030,2,6),date(2030,4,13),date(2030,4,14),date(2030,4,15),date(2030,4,16)],
    }
    for tatil in dini_tatiller.get(yil, []):
        tatiller.add(tatil)
    return tatiller


def is_gunu_mu(tarih):
    if tarih.weekday() >= 5:
        return False
    if tarih in turkiye_resmi_tatilleri(tarih.year):
        return False
    return True


def ay_is_gunu_bilgisi(referans_tarih):
    ay_ilk_gun = date(referans_tarih.year, referans_tarih.month, 1)
    ay_son_gun = date(referans_tarih.year, referans_tarih.month,
                      calendar.monthrange(referans_tarih.year, referans_tarih.month)[1])
    toplam_is_gunu = 0
    kalan_is_gunu = 0
    gun = ay_ilk_gun
    while gun <= ay_son_gun:
        if is_gunu_mu(gun):
            toplam_is_gunu += 1
            if gun >= referans_tarih:
                kalan_is_gunu += 1
        gun += timedelta(days=1)
    toplam_is_gunu = max(toplam_is_gunu, 1)
    kalan_is_gunu = max(kalan_is_gunu, 1)
    return toplam_is_gunu, kalan_is_gunu, ay_son_gun


def normalize_urun_adi(text):
    if pd.isna(text):
        return ""
    text = str(text)
    text = text.replace("l", "I")
    text = text.upper()
    replacements = {"İ": "I", "İ": "I", "ı": "I", "Ğ": "G", "Ü": "U", "Ş": "S", "Ö": "O", "Ç": "C"}
    for old, new in replacements.items():
        text = text.replace(old, new)
    for ch in [".", ",", ";", ":", "/", "\\", "-", "_", "(", ")", "[", "]", "*"]:
        text = text.replace(ch, "")
    text = re.sub(r"\s+", "", text)
    text = text.replace("%O", "%0")
    for _ in range(5):
        text = re.sub(r"(\d)O", r"\g<1>0", text)
    return text


LISTE_DISI_BASLANGICLAR = [normalize_urun_adi(x) for x in LISTE_DISI_BASLANGICLAR_RAW]


def liste_disi_baslangic_bul(normalize_key):
    normalize_key = str(normalize_key)
    for raw, key in zip(LISTE_DISI_BASLANGICLAR_RAW, LISTE_DISI_BASLANGICLAR):
        if normalize_key.startswith(key):
            return True, raw
    return False, ""


def gorunen_urun_adi_olustur(normalize_key):
    text = str(normalize_key)
    if text == "" or text == "NAN":
        return ""
    text = re.sub(r"%0(\d{2})(?=[A-Z])", r" %0.\1 ", text)
    text = re.sub(r"%(\d+(?:\.\d+)?)(?=[A-Z])", r" %\1 ", text)
    birimler = ["MCG", "MG", "ML", "CM", "MM", "CC", "IU", "GR"]
    birim_regex = "|".join(sorted(birimler, key=len, reverse=True))
    text = re.sub(rf"(\d+)({birim_regex})(?=[A-Z]|$)", r"\1 \2 ", text)
    text = re.sub(r"(?<=[A-Z])(?=\d)", " ", text)
    text = re.sub(r"(?<=\d)(?=[A-Z])", " ", text)
    kelimeler = sorted(set([
        "TABLET","KAPSUL","KAPSEL","KAPLET","FILM","FLAKON","AMPUL","SASE","SASET","DAMLA",
        "SURUP","SUSPANSIYON","SUSP","SOLUSYON","KREM","MERHEM","JEL","SPREY","PASTIL","PED",
        "SAFT","KASE","ADET","ORAL","NAZAL","GOZ","KULAK","BURUN","DERI","CILT","UZATILMIS",
        "SALIMLI","ENTERIK","KAPLI","RETARD","FORT","FORTE","PLUS","PEDIATRIK","YETISKIN",
        "COCUK","BEBEK","AROMALI","MUZ","CILEK","PORTAKAL","ORMAN","MEYVE","ENJEKTOR","INSULIN",
        "SUPER","KOMPRES","GAZ","STERIL","ENERJI","ENERGY","COZELTI",
    ]), key=len, reverse=True)
    for kelime in kelimeler:
        text = re.sub(rf"(?<!^)(?<!\s)({kelime})", r" \1", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def to_number(value):
    if pd.isna(value):
        return 0
    if isinstance(value, (int, float)):
        try:
            if math.isnan(value):
                return 0
        except Exception:
            pass
        return float(value)
    value = str(value).strip()
    if value == "":
        return 0
    value = value.replace(",", ".")
    try:
        return float(value)
    except Exception:
        return 0


def sayi_formatla_str(value):
    if isinstance(value, (int, float)):
        value = round(float(value), 2)
        if value == int(value):
            return str(int(value))
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value)


def siparis_durumu_belirle(row):
    ortalama_satis = row["ortalama_satis"]
    ham_siparis = row["ham_siparis_miktari"]
    hesap_stok = row["hesap_stok"]
    if (ortalama_satis > 10 and ham_siparis > 0
            and hesap_stok < (ortalama_satis / 30) * 7
            and hesap_stok <= 30):
        return "ACİL"
    if ham_siparis > 0:
        return "SİPARİŞ"
    return "GEREK YOK"


def siparis_hesapla(file_bytes: bytes):
    try:
        df = pd.read_excel(BytesIO(file_bytes), usecols="A,B,F")
    except Exception as e:
        raise ValueError("Dosya okunamadı. Eczanem Ürün Bazında Satış raporunu Excel formatında yükleyin.") from e

    df.columns = ["urun_adi", "toplam_3ay_satis", "stok"]
    df["normalize_ad"] = df["urun_adi"].apply(normalize_urun_adi)
    df["toplam_3ay_satis"] = df["toplam_3ay_satis"].apply(to_number)
    df["stok"] = df["stok"].apply(to_number)
    df = df[df["normalize_ad"] != ""]
    df = df[df["normalize_ad"] != "NAN"]

    sonuc = (
        df.groupby("normalize_ad", as_index=False)
        .agg(toplam_3ay_satis=("toplam_3ay_satis","sum"), stok=("stok","first"), ornek_ubs_adi=("urun_adi","first"))
    )

    bugun = bugun_turkiye()
    toplam_is_gunu, kalan_is_gunu, ay_son_gun = ay_is_gunu_bilgisi(bugun)

    sonuc["gorunen_urun_adi"] = sonuc["normalize_ad"].apply(gorunen_urun_adi_olustur)

    liste_disi_sonuclari = sonuc["normalize_ad"].apply(liste_disi_baslangic_bul)
    sonuc["liste_disi"] = liste_disi_sonuclari.apply(lambda x: x[0])
    sonuc["liste_disi_sebebi"] = liste_disi_sonuclari.apply(lambda x: f"Başlangıç: {x[1]}" if x[0] else "")

    haric = sonuc[sonuc["liste_disi"] == True].copy()
    sonuc = sonuc[sonuc["liste_disi"] == False].copy()

    # Yaz ayı kontrolü (Haziran=6, Temmuz=7, Ağustos=8)
    YAZ_AYLARI = {6, 7, 8}
    YAZ_INDIRIMI = 0.20
    yaz_ayi_mi = bugun.month in YAZ_AYLARI

    STOK_CAP = 50

    sonuc["ortalama_satis"] = (sonuc["toplam_3ay_satis"] / 3).round(2)
    sonuc["hesap_stok"] = sonuc["stok"].apply(lambda x: max(x, 0))
    sonuc["ortalama_gunluk_satis"] = (sonuc["ortalama_satis"] / 22).round(4)
    sonuc["kalan_ay_ihtiyaci"] = (sonuc["ortalama_gunluk_satis"] * kalan_is_gunu).round(2)
    sonuc["ham_siparis_miktari"] = sonuc.apply(
        lambda row: max(0, math.ceil(row["kalan_ay_ihtiyaci"] - row["hesap_stok"])), axis=1
    )

    # Yaz ayındaysa toplam sipariş miktarına %20 indirim uygula
    if yaz_ayi_mi:
        sonuc["ham_siparis_miktari"] = sonuc["ham_siparis_miktari"].apply(
            lambda x: max(0, math.ceil(x * (1 - YAZ_INDIRIMI))) if x > 0 else 0
        )

    # Parti sipariş = MIN(50, ham sipariş) — raf kapasitesi 50 ile sınırlı
    sonuc["planlanan_siparis_miktari"] = sonuc["ham_siparis_miktari"].apply(
        lambda x: min(STOK_CAP, x) if x > 0 else 0
    )
    sonuc["siparis_durumu"] = sonuc.apply(siparis_durumu_belirle, axis=1)
    sonuc["siparis_onceligi"] = sonuc["siparis_durumu"].map({"ACİL": 1, "SİPARİŞ": 2, "GEREK YOK": 3}).fillna(99)

    sonuc = sonuc.sort_values(
        by=["siparis_onceligi", "ortalama_satis"],
        ascending=[True, False]
    ).reset_index(drop=True)

    return sonuc, haric, bugun, toplam_is_gunu, kalan_is_gunu, ay_son_gun


def df_to_excel_bytes(sonuc: pd.DataFrame, haric: pd.DataFrame) -> bytes:
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    cols_map = {
        "gorunen_urun_adi": "Ürün Adı",
        "planlanan_siparis_miktari": "Parti Sipariş Mik.",
        "ham_siparis_miktari": "Top. Sipariş Mik.",
        "toplam_3ay_satis": "3 Aylık Satış",
        "ortalama_satis": "Ort. Aylık Satış",
        "stok": "Stok",
        "siparis_durumu": "Durum",
    }
    export_cols = [c for c in cols_map if c in sonuc.columns]
    df_out = sonuc[export_cols].rename(columns=cols_map)

    haric_cols = {
        "gorunen_urun_adi": "Ürün Adı",
        "toplam_3ay_satis": "3 Aylık Satış",
        "stok": "Stok",
        "liste_disi_sebebi": "Liste Dışı Sebebi",
    }
    haric_export = [c for c in haric_cols if c in haric.columns]
    haric_out = haric[haric_export].rename(columns=haric_cols) if haric_export else pd.DataFrame()

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Sipariş Sonuç")
        if not haric_out.empty:
            haric_out.to_excel(writer, index=False, sheet_name="Listeden Çıkarılanlar")

        for sheet_name, ws in writer.sheets.items():
            # Header stili
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Hücre hizalama
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            # Sütun genişlikleri
            for col_cells in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 45)
            ws.freeze_panes = "A2"

        # Renklendirme sadece Sipariş Sonuç sayfasında
        ws = writer.sheets["Sipariş Sonuç"]
        durum_col = None
        for cell in ws[1]:
            if cell.value == "Durum":
                durum_col = cell.column
                break
        if durum_col:
            acil_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            gerek_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
            for r in range(2, ws.max_row + 1):
                durum = ws.cell(row=r, column=durum_col).value
                fill = acil_fill if durum == "ACİL" else gerek_fill if durum == "GEREK YOK" else None
                if fill:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = fill

    return output.getvalue()


def df_to_pdf_bytes(sonuc: pd.DataFrame) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase import pdfmetrics
    from xml.sax.saxutils import escape

    output = BytesIO()
    font_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "C:/Windows/Fonts/arial.ttf",
        "DejaVuSans.ttf",
    ]
    font_name = "Helvetica"
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("TRFont", fp))
                font_name = "TRFont"
                break
            except Exception:
                pass

    pdf = SimpleDocTemplate(output, pagesize=A4,
                            rightMargin=0.8*cm, leftMargin=0.8*cm,
                            topMargin=0.8*cm, bottomMargin=0.8*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("T", parent=styles["Title"], fontName=font_name, fontSize=15, alignment=1, spaceAfter=8)
    normal_style = ParagraphStyle("N", parent=styles["Normal"], fontName=font_name, fontSize=7.2, leading=8.6)
    header_style = ParagraphStyle("H", parent=styles["Normal"], fontName=font_name, fontSize=7.2, leading=8.6, alignment=1)

    pdf_df = sonuc[sonuc["siparis_durumu"] == "ACİL"].copy()
    elements = [
        Paragraph("ACİL SİPARİŞ LİSTESİ", title_style),
        Paragraph(f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Acil Ürün Sayısı: {len(pdf_df)}", normal_style),
        Spacer(1, 8),
    ]

    def fmt(v):
        try:
            v = round(float(v), 2)
            return str(int(v)) if v == int(v) else f"{v:.2f}".rstrip("0").rstrip(".")
        except Exception:
            return str(v)

    table_data = [[
        Paragraph("Ürün Adı", header_style),
        Paragraph("Parti Sip. Mik.", header_style),
        Paragraph("Top. Sip. Mik.", header_style),
        Paragraph("Ort. Aylık", header_style),
        Paragraph("Stok", header_style),
    ]]
    for _, row in pdf_df.iterrows():
        table_data.append([
            Paragraph(escape(str(row.get("gorunen_urun_adi", ""))), normal_style),
            fmt(row.get("planlanan_siparis_miktari", 0)),
            fmt(row.get("ham_siparis_miktari", 0)),
            fmt(row.get("ortalama_satis", 0)),
            fmt(row.get("stok", 0)),
        ])

    tbl = Table(table_data, repeatRows=1, colWidths=[8.7*cm, 2.7*cm, 2.7*cm, 2.4*cm, 1.7*cm])
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), font_name),
        ("FONTSIZE", (0,0), (-1,-1), 7.2),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("GRID", (0,0), (-1,-1), 0.25, colors.black),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (1,1), (-1,-1), "CENTER"),
        ("ALIGN", (0,1), (0,-1), "LEFT"),
        ("LEFTPADDING", (0,0), (-1,-1), 3),
        ("RIGHTPADDING", (0,0), (-1,-1), 3),
        ("TOPPADDING", (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
    ]))
    elements.append(tbl)

    def sayfa_no(canvas, doc):
        canvas.saveState()
        canvas.setFont(font_name, 8)
        canvas.drawRightString(20.0*cm, 0.5*cm, f"Sayfa {doc.page}")
        canvas.restoreState()

    pdf.build(elements, onFirstPage=sayfa_no, onLaterPages=sayfa_no)
    return output.getvalue()


# =========================================================
# ENDPOINTS
# =========================================================

@app.get("/api/info")
def get_info():
    bugun = bugun_turkiye()
    toplam_is_gunu, kalan_is_gunu, ay_son_gun = ay_is_gunu_bilgisi(bugun)
    rapor_baslangic, rapor_bitis = onerilen_rapor_araligi()
    return {
        "bugun": bugun.isoformat(),
        "bugun_str": turkce_tarih_yaz(bugun),
        "aktif_ay": TURKCE_AYLAR[bugun.month],
        "toplam_is_gunu": toplam_is_gunu,
        "kalan_is_gunu": kalan_is_gunu,
        "ay_son_gun": ay_son_gun.isoformat(),
        "rapor_baslangic": rapor_baslangic.isoformat(),
        "rapor_bitis": rapor_bitis.isoformat(),
        "rapor_araligi_str": f"{turkce_tarih_yaz(rapor_baslangic)} - {turkce_tarih_yaz(rapor_bitis)}",
    }


@app.post("/api/hesapla")
async def hesapla(file: UploadFile = File(...)):
    if not file.filename.endswith((".xls", ".xlsx")):
        raise HTTPException(status_code=400, detail="Sadece .xls veya .xlsx dosyası yüklenebilir.")

    file_bytes = await file.read()

    try:
        sonuc, haric, bugun, toplam_is_gunu, kalan_is_gunu, ay_son_gun = siparis_hesapla(file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Hesaplama hatası: {str(e)}")

    def fmt(v):
        if isinstance(v, float):
            v = round(v, 2)
            if v == int(v):
                return int(v)
        return v

    cols_out = ["gorunen_urun_adi", "planlanan_siparis_miktari", "ham_siparis_miktari",
                "toplam_3ay_satis", "ortalama_satis", "stok", "siparis_durumu"]
    cols_out = [c for c in cols_out if c in sonuc.columns]
    sonuc_records = sonuc[cols_out].rename(columns={
        "gorunen_urun_adi": "urun_adi",
        "planlanan_siparis_miktari": "parti_siparis",
        "ham_siparis_miktari": "toplam_siparis",
        "parca_sayisi": "parti_sayisi",
        "toplam_3ay_satis": "satis_3ay",
        "ortalama_satis": "ort_aylik",
        "siparis_durumu": "durum",
    }).to_dict(orient="records")

    haric_cols = ["gorunen_urun_adi", "toplam_3ay_satis", "stok", "liste_disi_sebebi"]
    haric_cols = [c for c in haric_cols if c in haric.columns]
    haric_records = haric[haric_cols].rename(columns={
        "gorunen_urun_adi": "urun_adi",
        "toplam_3ay_satis": "satis_3ay",
        "liste_disi_sebebi": "sebep",
    }).to_dict(orient="records") if haric_cols else []

    return {
        "urunler": sonuc_records,
        "haric_tutulanlar": haric_records,
        "ozet": {
            "acil": int((sonuc["siparis_durumu"] == "ACİL").sum()),
            "siparis": int((sonuc["siparis_durumu"] == "SİPARİŞ").sum()),
            "gerek_yok": int((sonuc["siparis_durumu"] == "GEREK YOK").sum()),
            "liste_disi": len(haric),
            "kalan_is_gunu": kalan_is_gunu,
            "yaz_ayi_indirimi": bugun.month in {6, 7, 8},
        }
    }


@app.post("/api/excel-indir")
async def excel_indir(file: UploadFile = File(...)):
    file_bytes = await file.read()
    try:
        sonuc, haric, *_ = siparis_hesapla(file_bytes)
        excel_bytes = df_to_excel_bytes(sonuc, haric)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    return StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=siparis_sonuc.xlsx"}
    )


@app.post("/api/pdf-indir")
async def pdf_indir(file: UploadFile = File(...)):
    file_bytes = await file.read()
    try:
        sonuc, haric, *_ = siparis_hesapla(file_bytes)
        pdf_bytes = df_to_pdf_bytes(sonuc)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=acil_siparis_listesi.pdf"}
    )
